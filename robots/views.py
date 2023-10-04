from django.db.models import Count
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from openpyxl.workbook import Workbook
from robots.models import Robot


@csrf_exempt
def get_excel_report(request):
    model_data = Robot.objects.values('model', 'version').annotate(robot_count=Count('id'))

    if not model_data:
        return JsonResponse({'error': 'No data found'}, status=404)

    wb = Workbook()

    for entry in model_data:
        model = entry['model']
        version = entry['version']
        robot_count = entry['robot_count']

        if model not in wb.sheetnames:
            sheet = wb.create_sheet(model)
            sheet.append(["Модель", "Версия", "Количество за неделю"])
        else:
            sheet = wb[model]

        sheet.append([model, version, robot_count])

    default_sheet = wb.get_sheet_by_name('Sheet')
    wb.remove_sheet(default_sheet)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=robots_data.xlsx'
    wb.save(response)

    return response
