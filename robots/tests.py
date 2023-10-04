import io
from openpyxl import load_workbook
from django.urls import reverse
import pytest
from robots.models import Robot


@pytest.mark.django_db
def test_get_excel_report(client):
    Robot.objects.create(model='T1', version='V1')
    Robot.objects.create(model='T2', version='V2')

    response = client.get(reverse('report'))

    assert response.status_code == 200

    content = io.BytesIO(response.content)
    wb = load_workbook(content)

    assert 'T1' in wb.sheetnames

    sheet_t1 = wb['T1']
    assert sheet_t1.cell(row=1, column=1).value == 'Модель'
    assert sheet_t1.cell(row=1, column=2).value == 'Версия'
    assert sheet_t1.cell(row=1, column=3).value == 'Количество за неделю'
    assert sheet_t1.cell(row=2, column=1).value == 'T1'
    assert sheet_t1.cell(row=2, column=2).value == 'V1'
    assert sheet_t1.cell(row=2, column=3).value == 1

    assert 'T2' in wb.sheetnames

    sheet_t2 = wb['T2']
    assert sheet_t2.cell(row=1, column=1).value == 'Модель'
    assert sheet_t2.cell(row=1, column=2).value == 'Версия'
    assert sheet_t2.cell(row=1, column=3).value == 'Количество за неделю'
    assert sheet_t2.cell(row=2, column=1).value == 'T2'
    assert sheet_t2.cell(row=2, column=2).value == 'V2'
    assert sheet_t2.cell(row=2, column=3).value == 1
